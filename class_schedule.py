# -*- coding: utf-8 -*-

from openpyxl import Workbook

from openpyxl.styles import (
    Alignment,
    Font,
    Border,
    Side,
    PatternFill,
)

from progression_time import hour_progression
from file_reader import read_data


wb = Workbook()
dest_filename = "/home/carlos-machado/Documents/github/make_class_schedule/teste.xlsx"

ws1 = wb.active
ws1.title = "Schedule"

days = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]
columns = "BCDEFGH"
horarios = hour_progression(1)

#
# Setting up our styles that must be used
#

# head alignment/font [from days of week]
head_font = Font(name="Times New Roman", size=16, bold=True, color="CECAC5")
general_alignment = Alignment(horizontal="center", vertical="center")

# time font
time_font = Font(name="Times New Roman", size=13, bold=True, color="CECAC5")

# subjects font
data_font = Font(name="Times New Roman", size=12, color="CECAC5")

# border
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
    )

# foreground color for head/time
headers_fgc = PatternFill(fill_type="solid", fgColor="2C001E")

# foreground color for data[subjects]
data_fgc = PatternFill(fill_type="solid", fgColor="AEA79F")

#
# Styles session finished
#

for i in range(1, 9):
    for j in range(1, 17):
        cell = ws1.cell(row=j, column=i)
        cell.fill = headers_fgc
        cell.border = border

# define our head values and styles
ws1.column_dimensions["A"].width = 12.2
for col in range(2, 9):
    head = ws1.cell(row=1, column=col, value=days[col - 2])
    head.font = head_font
    head.alignment = general_alignment
    ws1.column_dimensions[columns[col - 2]].width = 18.24

# define our time values and styles
ws1.row_dimensions[1].height = 22.4
for row_ in range(2, len(horarios) + 2):
    time = ws1.cell(row=row_, column=1, value=horarios[row_ - 2])
    time.font = time_font
    ws1.row_dimensions[row_].height = 22.4

data = read_data("file_example.txt")


wb.save(filename=dest_filename)
